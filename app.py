from flask import Flask, render_template, request, redirect, url_for, send_from_directory
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)

# Directory to store uploaded files
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)  # Create the directory if it doesn't exist

# Directory to store processed files
PROCESSED_FOLDER = 'processed'
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'xls', 'xlsx'}  # Allowed file extensions

def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Step 1: Process initial Excel file
def process_initial_excel(input_file, output_file):
    try:
        dfhead = pd.read_excel(input_file, skiprows=16)
        var_name = dfhead.iloc[0].dropna().values
        var_data = dfhead.iloc[1].dropna().values
        des = []
        for i in range(min(len(var_name), len(var_data))):
            name = var_name[i]
            data = var_data[i]
            des.append([name, data])
        brand_value = next((d[1] for d in des if d[0] == 'Brand'), '')
        copyline_value = next((d[1] for d in des if d[0] == 'Copyline'), '')
        df = pd.read_excel(input_file, skiprows=19)
        summary_stop_index = df[df.iloc[:, 0].astype(str).str.contains("Summary for Columns", case=False, na=False)].index.min()
        summary_rows = df.iloc[:, 0].astype(str).str.contains("summary", case=False, na=False)
        valid_rows = (~summary_rows) & (df.index <= summary_stop_index)
        df.loc[valid_rows] = df.loc[valid_rows].fillna(method='ffill')
        df = df.iloc[:summary_stop_index + 1]
        df.to_excel(output_file, index=False)
        return brand_value, copyline_value
    except Exception as e:
        raise Exception(f"Error in process_initial_excel: {e}") # Re-raise the exception with context

# Step 2: Transform data by Channel
def sanitize_sheet_name(name):
    return re.sub(r'[\\/*?:[\]]', '', str(name))[:31]

    
def normalize_datetime(row):
    """
    Normalize datetime when time is over 24:00
    Returns tuple of (day_of_week, date, start_time)
    """
    # แปลง string เป็น datetime โดยใช้ dayfirst=True เพื่อให้วันที่ถูกต้องตามรูปแบบ DD/MM/YYYY
    date = pd.to_datetime(row['Date'], errors='coerce', dayfirst=True)
    if pd.isna(date):
        return None, None, None

    time_str = row['Start Time']
    
    # แยกชั่วโมงและนาที
    hours, minutes = map(int, time_str.split(':'))
    
    # ถ้าชั่วโมงมากกว่าหรือเท่ากับ 24
    if hours >= 24:
        # ปรับชั่วโมงและเพิ่มวัน
        hours -= 24
        date = date + pd.Timedelta(days=1)
        
    # สร้างเวลาใหม่ในรูปแบบ HH:MM
    time_str = f"{hours:02d}:{minutes:02d}"
    
    # อัพเดท Day of Week
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    day_of_week = days[date.weekday()]
    
    return day_of_week, date, time_str

def transform_excel_by_channel(input_file, output_file, brand_value, copyline_value):
    try:
        xlsx = pd.ExcelFile(input_file, engine="openpyxl")
        total_rows = 0
        is_first_sheet = True

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            last_sheet_name = None
            last_sheet_rows = 0
            
            for sheet_name in xlsx.sheet_names:
                df = pd.read_excel(input_file, sheet_name=sheet_name, engine="openpyxl")
                df['Duration'] = df['Duration'].apply(lambda x: str(x).split(':')[-1] if pd.notna(x) else x)
                df['No. Of Spots'] = ''
                df = df[~df['Channel'].str.contains('Summary', na=False)]
                df = df.dropna(subset=['Channel'])
                
                channels = df['Channel'].unique()
                for channel in channels:
                    channel_df = df[df['Channel'] == channel].copy()
                    
                    # แปลงและจัดการวันที่/เวลา
                    normalized_data = channel_df.apply(normalize_datetime, axis=1)
                    channel_df['Day Of Week'] = normalized_data.apply(lambda x: x[0])
                    channel_df['Date'] = normalized_data.apply(lambda x: x[1])
                    channel_df['Start Time'] = normalized_data.apply(lambda x: x[2])
                    
                    # เรียงข้อมูลตามวันที่และเวลา
                    channel_df = channel_df.dropna(subset=['Date', 'Start Time'])
                    channel_df = channel_df.sort_values(['Date', 'Start Time'])
                    
                    # แปลงวันที่กลับเป็น string format (DD/MM/YYYY)
                    channel_df['Date'] = channel_df['Date'].dt.strftime('%d/%m/%Y')
                    
                    columns = channel_df.columns.tolist()
                    columns.remove('Channel')

                    # สร้าง new_data ตามรูปแบบที่ต้องการ
                    new_data = []
                    
                    if is_first_sheet:
                        new_data.append([brand_value, '', '', '', '', '',copyline_value])
                        is_first_sheet = False

                    new_data.append([channel, f'Brand : {brand_value}', '', '', '', '',copyline_value])
                    new_data.append(columns)
                    new_data.extend(channel_df[columns].values.tolist())

                    summary_row = [''] * len(columns)
                    summary_row[0] = channel
                    summary_row[1] = f'Brand : {brand_value}'
                    summary_row[6] = f'Total {len(channel_df)} Spots'
                    new_data.append(summary_row)

                    safe_sheet_name = sanitize_sheet_name(channel)
                    pd.DataFrame(new_data).to_excel(writer, sheet_name=safe_sheet_name, index=False, header=False)
                    
                    total_rows += len(channel_df)
                    last_sheet_name = safe_sheet_name
                    last_sheet_rows = len(new_data)

            if last_sheet_name:
                grand_total_row = [''] * len(columns)
                grand_total_row[0] = f'Brand : {brand_value}'
                grand_total_row[6] = f'Grand Total {total_rows} Spots'
                pd.DataFrame([grand_total_row]).to_excel(writer, sheet_name=last_sheet_name, index=False, header=False, startrow=last_sheet_rows)
    except Exception as e:
        raise Exception(f"Error in transform_excel_by_channel: {e}")
    
def format_excel(input_file, output_file):
    try:
        excel_file = pd.ExcelFile(input_file, engine="openpyxl")
        dfs = []
        is_first_sheet = True
        for sheet_name in excel_file.sheet_names:
            df = excel_file.parse(sheet_name, header=None)
            if is_first_sheet:
                df.iloc[2, 0] = ""
                df.iloc[2, 1] = "Date/Time"
                df.iloc[2, 2] = ""
                df.iloc[2, 3] = "Brk"
                df.iloc[2, 4] = "PIB"
                df.iloc[2, 5] = "Dur"
                df.iloc[2, 6] = "Program"
                df.iloc[2, 7] = "Remark"
                is_first_sheet = False
            else:
                df.iloc[1, 0] = ""
                df.iloc[1, 1] = "Date/Time"
                df.iloc[1, 2] = ""
                df.iloc[1, 3] = "Brk"
                df.iloc[1, 4] = "PIB"
                df.iloc[1, 5] = "Dur"
                df.iloc[1, 6] = "Program"
                df.iloc[1, 7] = "Remark"
            dfs.append(df)

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            startrow = 0
            for df in dfs:
                df.to_excel(writer, sheet_name="Sheet1", startrow=startrow, index=False, header=False)
                startrow += len(df) + 1

        wb = load_workbook(output_file)
        ws = wb["Sheet1"]
        ws.insert_rows(1, amount=3)
        font_24 = Font(size=20)
        font_7 = Font(size=6)
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells('A1:B3')
        ws.merge_cells('C1:G2')
        ws['C1'] = "The Nielsen Company (Thailand) Limited."
        ws['C1'].font = font_24
        ws['C1'].alignment = alignment_center

        ws['H1'] = "Daily Comercial Logs"
        ws['H1'].font = font_7
        ws['H1'].alignment = alignment_center
        ws['H2'] = "Advertisment Activity"
        ws['H2'].font = font_7
        ws['H2'].alignment = alignment_center
        ws['H3'] = "By Copyline"
        ws['H3'].font = font_7
        ws['H3'].alignment = alignment_center

        ws.merge_cells('C3:G3')
        ws['C3'] = "34th Fls., United Center, 323 Silom Rd..Bangkok 10500 Tel. 0-2674-6000 Fax. 0-274-6000 Ext.5102"
        ws['C3'].font = font_7
        ws['C3'].alignment = alignment_center

        column_widths = {"A": 8, "B": 8, "C": 6, "D": 6, "E": 6, "F": 6, "G": 40, "H": 8}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        highlight_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        highlight_font = Font(bold=True, size=7)
        normal_font = Font(bold=False, size=7)

        for row in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=8):
            for cell in row:
                cell.font = highlight_font

        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=8):
            is_brand_row = False
            is_date_time_row = False
            for cell in row:
                if cell.value and "Brand :" in str(cell.value):
                    is_brand_row = True
                    break
                if cell.value and "Date/Time" in str(cell.value):  # ถ้ามีคำว่า "Date/Time"
                    is_date_time_row = True
                    break
            for col in row:
                if is_brand_row:  # ถ้าเจอ "Brand :"
                    col.fill = highlight_fill  # ลงสีพื้น
                    col.font = highlight_font  # ฟอนต์ตัวหนา ขนาด 7
                elif is_date_time_row:  # ถ้าเจอ "Date/Time"
                    col.font = highlight_font  # ฟอนต์ตัวหนา ขนาด 7 (แต่ไม่ลงสี)
                else:
                    col.font = normal_font  # ฟอนต์ปกติ ขนาด 6

        border_top_bottom = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
        for row in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=8):
            for cell in row:
                cell.border = border_top_bottom

        wb.save(output_file)
    except Exception as e:
        raise Exception(f"Error in format_excel: {e}")


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = file.filename
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            intermediate_file = os.path.join(app.config['UPLOAD_FOLDER'], "unmerge.xlsx")
            transformed_file = os.path.join(app.config['PROCESSED_FOLDER'], "transformed_data.xlsx")
            final_output = os.path.join(app.config['PROCESSED_FOLDER'], "format.xlsx")

            try:
                brand_value, copyline_value = process_initial_excel(filepath, intermediate_file)
                transform_excel_by_channel(intermediate_file, transformed_file, brand_value, copyline_value)
                format_excel(transformed_file, final_output)

                return send_from_directory(app.config['PROCESSED_FOLDER'], "format.xlsx", as_attachment=True)

            except Exception as e:  # Catch any errors during processing
                return f"An error occurred during processing: {str(e)}"  # Display error message

        else:
            return "Invalid file type. Please upload an .xls or .xlsx file."

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
